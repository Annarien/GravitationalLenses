"""
This makes an excel spreadsheet to make the the results from ml_Lenses more readible, instead of writing it in the terminal.
"""
import astropy.table  as atpy
import numpy as np
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

def makeInitialTable():
    
    tab = atpy.Table()
    tab.add_column(atpy.Column(np.zeros(1), "Test"))
    tab.add_column(atpy.Column(np.zeros(1), "Description"))
    tab.add_column(atpy.Column(np.zeros(1), "imageTrain_std"))
    tab.add_column(atpy.Column(np.zeros(1), "imageTrain_mean"))
    tab.add_column(atpy.Column(np.zeros(1), "imageTrain_shape"))
    tab.add_column(atpy.Column(np.zeros(1), "imageLabels_shape"))
    tab.add_column(atpy.Column(np.zeros(1), "train_percent"))
    tab.add_column(atpy.Column(np.zeros(1), "test_percent"))
    tab.add_column(atpy.Column(np.zeros(1), "xTrain_shape"))
    tab.add_column(atpy.Column(np.zeros(1), "xTest_shape"))
    tab.add_column(atpy.Column(np.zeros(1), "yTrain_shape"))
    tab.add_column(atpy.Column(np.zeros(1), "yTest_shape"))
    tab.add_column(atpy.Column(np.zeros(1), "n_splits"))
    tab.add_column(atpy.Column(np.zeros(1), "random_state"))
    tab.add_column(atpy.Column(np.zeros(1), "AccuracyScore"))
    tab.add_column(atpy.Column(np.zeros(1), "KFoldAccuracy"))
    tab.add_column(atpy.Column(np.zeros(1), "AccuracyScore_47"))
    tab.add_column(atpy.Column(np.zeros(1), "KFoldAccuracy_47"))
    tab.add_column(atpy.Column(np.zeros(1), "AccuracyScore_84"))
    tab.add_column(atpy.Column(np.zeros(1), "KFoldAccuracy_84"))
    tab.add_column(atpy.Column(np.zeros(1), "AccuracyScore_131"))
    tab.add_column(atpy.Column(np.zeros(1), "KFoldAccuracy_131"))
    tab.write('../Results/ml_Lenses_results.csv', overwrite = True )
    return (tab)

def getRowNum():
    # find out how many rows in table

    wb = load_workbook('ml_Lenses_results.xlsx', use_iterators=True)
    sheet = wb.worksheets[0]

    row_count = sheet.max_row
    column_count = sheet.max_column

    currentRow = row_count + 1

    return (currentRow)

def addRowToTable(tab, description, train_percent, test_percent, n_splits, random_state, imageTrain_std, imageTrain_mean, imageTrain_shape, imageLabels_shape, xTrain_shape, xTest_shape, yTrain_shape, yTest_shape, AccuracyScore, KFoldAccuracy, AccuracyScore_47, KFoldAccuracy_47, AccuracyScore_84, KFoldAccuracy_84, AccuracyScore_131, KFoldAccuracy_131):
    currentRow = getRowNum()

    tab["Test"][currentRow] = currentRow
    tab["Description"][currentRow] = description
    tab["imageTrain_std"][currentRow] = imageTrain_std
    tab["imageTrain_mean"][currentRow] = imageTrain_mean
    tab["imageTrain_shape"][currentRow] = imageTrain_shape
    tab["imageLabels_shape"][currentRow] = imageLabels_shape
    tab["train_percent"][currentRow] = train_percent
    tab["test_percent"][currentRow] = test_percent
    tab["xTrain_shape"][currentRow] = xTrain_shape
    tab["xTest_shape"][currentRow] = xTest_shape
    tab["yTrain_shape"][currentRow] = yTrain_shape
    tab["yTest_shape"][currentRow] = yTest_shape
    tab["n_splits"][currentRow] = n_splits
    tab["random_state"][currentRow] = random_state
    tab["AccuracyScore"][currentRow] = AccuracyScore
    tab["KFoldAccuracy"][currentRow] = KFoldAccuracy
    tab["AccuracyScore_47"][currentRow] = AccuracyScore_47
    tab["KFoldAccuracy_47"][currentRow] = KFoldAccuracy_47
    tab["AccuracyScore_84"][currentRow] = AccuracyScore_84
    tab["KFoldAccuracy_84"][currentRow] = KFoldAccuracy_84
    tab["AccuracyScore_131"][currentRow] = AccuracyScore_131
    tab["KFoldAccuracy_131"][currentRow] = KFoldAccuracy_131
    tab.write('../Results/ml_Lenses_results.csv', overwrite = True)

#_______________________________________________________________________________________________________
makeInitialTable()
